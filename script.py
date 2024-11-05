import socket
import openpyxl
from openpyxl import Workbook
import json
import os

def create_or_update_excel(data):
    # Ensure that the input is a dictionary and not empty
    if not data or not isinstance(data, dict):
        print("No data or invalid data provided.")
        return

    # Extracting the client name and preparing the file name
    client_name = data.get("client")    
    file_name = f"{client_name}.xlsx"
    
    # Check if the file exists
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hardware Data"
        headers = ["Type", "Model", "Serie", "Inventaire", "Created By"]
        ws.append(headers)

    existing_serials = {row[2] for row in ws.iter_rows(min_row=2, values_only=True)}

    serial = data.get("serie")
    if serial not in existing_serials:
        ws.append([data["type"], data["model"], serial, data["inventaire"], data["createdBy"]])
        existing_serials.add(serial)
    else:
        print(f"Duplicate entry found for serial: {serial}, not added.")

    # Save the workbook
    wb.save(file_name)
    print(f"Data has been saved to {file_name}")


# Create TCP/IP socket
server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

host = '0.0.0.0' # enter your ip here (default all interfaces)
port = 5000
server_socket.bind((host, port))

# Listen for incoming connections
server_socket.listen(5)
print(f"Server listening on {host}:{port}")

try:
    while True:
        connection, client_address = server_socket.accept()
        try:
            print(f"Connection from {client_address}")

            data = connection.recv(1024).decode()
            if data:
                print(f"Received data: {data}")
                json_data = json.loads(data)
                create_or_update_excel(json_data) 

                connection.sendall(b'Data received successfully!')
            else:
                print("No data received")
        finally:
            connection.close()
except KeyboardInterrupt:
    print("Server is shutting down...")
finally:
    server_socket.close()